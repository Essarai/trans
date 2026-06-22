"""
converter.py - 维普 Excel -> 浙大社 ZD_JATS XML 转换器

用法:
    from converter import convert, JournalMeta

    convert(
        JournalMeta(title_zh="机电工程", issn="1001-4551", cn="33-1088/TH", publisher="浙江大学"),
        "维普数据.xlsx",
        "PDF",          # PDF 根目录，None 则不复制
        "./output",
    )

    # 命令行
    python3 converter.py --xlsx 维普数据.xlsx --output-dir ./output \\
        --journal-title 机电工程 --issn 1001-4551 --cn 33-1088/TH --publisher 浙江大学

    python3 converter.py --with-pdf --pdf-dir ./PDF
    python3 converter.py --gui
"""

from __future__ import annotations

import argparse
import re
import shutil
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Mapping, Optional, Union

import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_XLSX = SCRIPT_DIR / "测试样本" / "维普数据.xlsx"
DEFAULT_PDF_DIR = SCRIPT_DIR / "测试样本" / "PDF"
DEFAULT_OUTPUT = SCRIPT_DIR.parent / "output"


# ============== 期刊元数据 ==============

@dataclass
class JournalMeta:
    title_zh: str = "机电工程"
    issn: str = "1001-4551"
    cn: str = "33-1088/TH"
    publisher: str = "浙江大学"


DEFAULT_JOURNAL_META = JournalMeta()

# Excel 可选列名 -> JournalMeta 字段
_JOURNAL_COL_MAP: dict[str, str] = {}
for _field, aliases in {
    "title_zh": ("期刊名", "期刊", "journal-title", "journal"),
    "issn": ("issn", "ISSN"),
    "cn": ("cn", "CN", "国内刊号"),
    "publisher": ("出版社", "publisher", "publisher-name"),
}.items():
    for alias in aliases:
        _JOURNAL_COL_MAP[alias] = _field

_ISSN_FROM_DOI_RE = re.compile(r"issn[\./](\d{4}-\d{3,4})", re.IGNORECASE)


def infer_issn_from_doi(doi: str) -> str:
    m = _ISSN_FROM_DOI_RE.search(doi or "")
    return m.group(1) if m else ""


def load_journal_meta_from_excel(xlsx_path: Path) -> JournalMeta:
    """从 Excel 可选列读取期刊元数据 (取首条非空值)."""
    meta = JournalMeta()
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows, ())]
    col_idx = {h: i for i, h in enumerate(header) if h}
    field_cols: dict[str, int] = {}
    for col_name, idx in col_idx.items():
        field = _JOURNAL_COL_MAP.get(col_name)
        if field and field not in field_cols:
            field_cols[field] = idx
    if not field_cols:
        wb.close()
        return meta
    for row in rows:
        if not row or all(c is None for c in row):
            continue
        for field, idx in field_cols.items():
            if idx < len(row) and row[idx]:
                setattr(meta, field, str(row[idx]).strip())
        break
    wb.close()
    return meta


def resolve_journal_meta(
    xlsx_path: Path,
    arts: list[Article],
    overrides: Optional[dict[str, str]] = None,
) -> JournalMeta:
    """合并默认值、Excel 可选列、命令行参数、DOI 推断 ISSN."""
    meta = load_journal_meta_from_excel(xlsx_path)
    if overrides:
        for key, val in overrides.items():
            if val:
                setattr(meta, key, val)
    if not meta.issn:
        for art in arts:
            inferred = infer_issn_from_doi(art.doi)
            if inferred:
                meta.issn = inferred
                break
    return meta


def normalize_journal(
    journal: Union[JournalMeta, Mapping[str, str], None],
) -> Optional[dict[str, str]]:
    """将期刊元数据统一为 overrides 字典 (空值字段省略)."""
    if journal is None:
        return None
    if isinstance(journal, JournalMeta):
        raw = {
            "title_zh": journal.title_zh,
            "issn": journal.issn,
            "cn": journal.cn,
            "publisher": journal.publisher,
        }
    else:
        raw = dict(journal)
    return {k: str(v).strip() for k, v in raw.items() if v and str(v).strip()}


def build_journal_meta(
    xlsx_path: Path,
    arts: list,
    journal: Union[JournalMeta, Mapping[str, str], None] = None,
) -> JournalMeta:
    """合并 Excel 可选列、传入的期刊元数据、DOI 推断 ISSN."""
    if isinstance(journal, JournalMeta):
        meta = journal
        if not meta.issn:
            for art in arts:
                inferred = infer_issn_from_doi(art.doi)
                if inferred:
                    meta.issn = inferred
                    break
        return meta
    overrides = normalize_journal(journal)
    return resolve_journal_meta(xlsx_path, arts, overrides)


# ============== 通用: XML 字符处理 ==============

def xml_escape(text) -> str:
    """XML 字符转义 """
    if text is None:
        return ""
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


_SUP_SUB_PATTERN = re.compile(
    r"\^\(([^)]+)\)"
    r"|_\(([^)]+)\)"
)
_EMAIL_ADDR_RE = re.compile(r"[\w.\-+]+@[\w.\-]+\.\w+")


def _protect_emails(text: str) -> tuple[str, list[str]]:
    emails: list[str] = []

    def repl(m: re.Match) -> str:
        emails.append(m.group(0))
        return f"\x00EMAIL{len(emails) - 1}\x00"

    return _EMAIL_ADDR_RE.sub(repl, text), emails


def _restore_emails(text: str, emails: list[str]) -> str:
    for i, email in enumerate(emails):
        text = text.replace(f"\x00EMAIL{i}\x00", email)
    return text


def convert_super_sub(text: Optional[str]) -> str:
    """维普 ^X / ^(xxx) -> <sup>; _X / _(xxx) -> <sub>.

    邮箱地址中的下划线不做上下标转换.
    """
    if not text:
        return ""
    protected, emails = _protect_emails(str(text))
    escaped = xml_escape(protected)

    def repl(m: re.Match) -> str:
        if m.group(1) is not None:
            return f"<sup>{m.group(1)}</sup>"
        if m.group(2) is not None:
            return f"<sub>{m.group(2)}</sub>"
        return m.group(0)

    return _restore_emails(_SUP_SUB_PATTERN.sub(repl, escaped), emails)


# ============== 解析函数 ==============

def parse_multi(value, seps: str = ",;；") -> list:
    if not value:
        return []
    return [v.strip() for v in re.split(f"[{re.escape(seps)}]", str(value)) if v.strip()]


def parse_semicolon(value) -> list:
    """按分号拆分 (作者/单位字段内含逗号, 不能用逗号作分隔符)."""
    if not value:
        return []
    return [v.strip() for v in re.split(r"[;；]", str(value)) if v.strip()]


def parse_authors(value) -> list:
    """中文作者: 'name1id1;name2id2;' -> [{name_zh, aff_ids}]."""
    out = []
    for raw in parse_semicolon(value):
        m = re.match(r"^(.+?)([\d,]+)$", raw)
        if m:
            name = m.group(1).strip()
            affs = [x for x in m.group(2).split(",") if x]
        else:
            name = raw
            affs = []
        out.append({"name_zh": name, "aff_ids": affs})
    return out


def parse_en_authors(value) -> list:
    """英文作者: 'SURNAME Given-name id1,id2;'.

    返回 {surname_en, given_en, aff_ids}.
    """
    out = []
    for raw in parse_semicolon(value):
        m = re.match(r"^(.+?)\s+([A-Za-z\-\.]+?)([\d,]*)$", raw)
        if m:
            surname = m.group(1).strip()
            given = m.group(2).strip()
            affs = [x for x in m.group(3).split(",") if x] if m.group(3) else []
        else:
            m2 = re.match(r"^(.+?)([\d,]+)$", raw)
            if m2:
                surname = m2.group(1).strip()
                given = ""
                affs = [x for x in m2.group(2).split(",") if x]
            else:
                surname = raw
                given = ""
                affs = []
        out.append({"surname_en": surname, "given_en": given, "aff_ids": affs})
    return out


def parse_affiliation_text(raw: str) -> tuple:
    """单位: 'Nname,city postal' -> (label, text).

    样例中单位以原始字符串作为 <aff> 文本内容, 不拆分 institution/city/postal.
    """
    raw = raw.strip()
    m = re.match(r"^(\d+)(.+)$", raw)
    if m:
        return m.group(1), m.group(2).strip()
    return "", raw


def parse_affiliations(value) -> list:
    return [
        {"id": label, "text": text}
        for label, text in (parse_affiliation_text(r) for r in parse_semicolon(value))
    ]


def parse_kw(value) -> list:
    return parse_multi(value, seps=",;；")


def parse_clc(value) -> list:
    """CLC 号拆分. 全部放在一个 <subj-group subj-group-type='clc'> 下."""
    return parse_multi(value, seps=",;；")


def parse_funding(value) -> tuple[list, str]:
    """基金: '项目名(num1,num2);项目名2(num3);' -> ([{name, ids}], funding-statement 原文)."""
    raw_all = str(value).strip() if value else ""
    out = []
    for raw in parse_multi(raw_all, seps=";；"):
        raw = raw.strip()
        if not raw:
            continue
        if "(" in raw:
            last_paren = raw.rfind("(")
            name = raw[:last_paren]
            id_part = raw[last_paren + 1 :].replace("(", "").replace(")", "")
            ids = [x.strip() for x in id_part.split(",") if x.strip()]
        else:
            name = raw
            ids = []
        out.append({"name": name, "ids": ids})
    return out, raw_all


def parse_dt(value) -> str:
    """日期统一输出 'YYYY-M-D' (不补零, 与样例一致)."""
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return f"{value.year}-{value.month}-{value.day}"
    s = str(value).strip()
    m = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", s)
    if m:
        return f"{m.group(1)}-{int(m.group(2))}-{int(m.group(3))}"
    return ""


_EMAIL_RE = re.compile(r"E-?mail\s*[:：]\s*([^\s。,，;；]+)", re.IGNORECASE)


def parse_bio(value) -> dict:
    """解析 '通讯作者' 列.

    形式 1 (first = corresp):
        "name(year-),...。E-mail:xxx。"
    形式 2 (first != corresp, 使用「通信联系人」):
        "first_bio。E-mail:a@x;通信联系人:corresp_bio。E-mail:b@y。"
    形式 3 (通讯联系人/通信作者 保留在作者简介内):
        "...;通信作者:...;..." / "...;通讯联系人:..."
    """
    empty = {
        "first_bio": "",
        "first_email": "",
        "corresp_bio": "",
        "corresp_email": "",
        "has_separate_corresp": False,
        "corresp_name_zh": "",
    }
    if not value:
        return empty
    text = str(value).strip()
    parts = re.split(r"[;；]\s*通信联系人[:：]\s*", text, maxsplit=1)
    first = parts[0].strip()
    corresp = parts[1].strip() if len(parts) > 1 else ""

    def extract_email(s: str) -> str:
        m = _EMAIL_RE.search(s)
        if not m:
            return ""
        return m.group(1).rstrip(".")

    first_email = extract_email(first)
    if corresp:
        corresp_email = extract_email(corresp)
        m_name = re.match(r"^([^,，。;；\s]+)", corresp)
        corresp_name_zh = m_name.group(1) if m_name else ""
        return {
            "first_bio": first,
            "first_email": first_email,
            "corresp_bio": corresp,
            "corresp_email": corresp_email,
            "has_separate_corresp": True,
            "corresp_name_zh": corresp_name_zh,
        }
    return {
        "first_bio": first,
        "first_email": first_email,
        "corresp_bio": "",
        "corresp_email": "",
        "has_separate_corresp": False,
        "corresp_name_zh": "",
    }


# ============== 数据模型 ==============

@dataclass
class Article:
    doi: str = ""
    title_zh: str = ""
    title_en: str = ""
    col_zh: str = ""
    col_en: str = ""
    clc: list = field(default_factory=list)
    abstract_zh: str = ""
    abstract_en: str = ""
    kw_zh: list = field(default_factory=list)
    kw_en: list = field(default_factory=list)
    authors_zh: list = field(default_factory=list)
    authors_en: list = field(default_factory=list)
    aff_zh: list = field(default_factory=list)
    aff_en: list = field(default_factory=list)
    pub_date: str = ""
    volume: str = ""
    issue: str = ""
    fpage: str = ""
    lpage: str = ""
    funding: list = field(default_factory=list)
    funding_statement: str = ""
    pdf_path: str = ""
    first_bio: str = ""
    first_email: str = ""
    corresp_bio: str = ""
    corresp_email: str = ""
    has_separate_corresp: bool = False
    corresp_name_zh: str = ""


# ============== XML 构造 (对齐 trans/example/ 样例) ==============

def _wrap_email_in_text(text: str, email: str = "") -> str:
    """把文本中所有 'E-mail: addr' 替换成 'E-mail: <email>addr</email>'."""

    def repl(m: re.Match) -> str:
        prefix = m.group(1)
        addr = m.group(2).rstrip(".")
        suffix = m.group(3) or ""
        return prefix.rstrip() + " <email>" + addr + "</email>" + suffix

    pattern = re.compile(
        r"(E-?mail\s*[:：]\s*)([\w.\-+]+@[\w.\-]+\.\w+)(\.)?",
        flags=re.IGNORECASE,
    )
    return pattern.sub(repl, text)


def _build_funding_xml(art: Article) -> list[str]:
    """构建 funding-group 片段 (紧凑格式, 与 example 一致)."""
    if not art.funding and not art.funding_statement:
        return []
    lines: list[str] = []
    lines.append("<funding-group>")

    if "%" in art.funding_statement and not re.search(r"[;；]", art.funding_statement):
        parts = art.funding_statement.split("%", 1)
        m1 = re.match(r"^(.+\()([^)]+)\)$", parts[1].strip()) if len(parts) > 1 else None
        if m1:
            source = parts[0] + "%" + m1.group(1)[:-1]
            award_id = m1.group(2)
            lines.append("<award-group>")
            lines.append(f"<funding-source>{xml_escape(source)}</funding-source>")
            lines.append(f"<award-id>{xml_escape(award_id)}</award-id></award-group>")
        else:
            for f in art.funding:
                lines.append("<award-group>")
                if f["name"]:
                    lines.append(f"<funding-source>{xml_escape(f['name'])}</funding-source>")
                for i, aid in enumerate(f["ids"]):
                    if i == len(f["ids"]) - 1:
                        lines.append(f"<award-id>{xml_escape(aid)}</award-id></award-group>")
                    else:
                        lines.append(f"<award-id>{xml_escape(aid)}</award-id>")
    else:
        for f in art.funding:
            lines.append("<award-group>")
            if f["name"]:
                lines.append(f"<funding-source>{xml_escape(f['name'])}</funding-source>")
            if f["ids"]:
                ids_xml = "".join(
                    f"<award-id>{xml_escape(aid)}</award-id>" for aid in f["ids"]
                )
                lines.append(f"{ids_xml}</award-group>")
            else:
                lines.append("</award-group>")

    if art.funding_statement:
        stmt = art.funding_statement.replace(";", "；")
        lines.append(f"<funding-statement>{xml_escape(stmt)}</funding-statement>")
    lines.append("</funding-group>")
    return lines


def build_article_xml(art: Article, journal: JournalMeta = DEFAULT_JOURNAL_META) -> str:
    p: list[str] = []
    p.append('\ufeff<?xml version="1.0" encoding="utf-8"?>')
    p.append(
        '<article dtd-version="1.3" xml:lang="zh" '
        'xmlns:mml="http://www.w3.org/1998/Math/MathML" '
        'xmlns:xlink="http://www.w3.org/1999/xlink" '
        'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" '
        'xsi:noNamespaceSchemaLocation="ZD_JATS-archivearticle1-3.xsd">'
    )
    p.append("<front>")

    # ---- <journal-meta> ----
    p.append("<journal-meta>")
    p.append("<journal-title-group>")
    p.append(f"<journal-title>{xml_escape(journal.title_zh)}</journal-title>")
    p.append("</journal-title-group>")
    p.append(f'<issn publication-format="print">{xml_escape(journal.issn)}</issn>')
    p.append(f'<cn publication-format="print">{xml_escape(journal.cn)}</cn>')
    p.append("<publisher>")
    p.append(f"<publisher-name>{xml_escape(journal.publisher)}</publisher-name>")
    p.append("</publisher>")
    p.append("</journal-meta>")

    # ---- <article-meta> ----
    p.append("<article-meta>")
    if art.doi:
        p.append(f'<article-id pub-id-type="doi">{xml_escape(art.doi)}</article-id>')

    # <article-categories>
    if art.col_zh or art.col_en or art.clc:
        p.append("<article-categories>")
        if art.col_zh:
            p.append('<subj-group subj-group-type="heading">')
            p.append(f"<subject>{xml_escape(art.col_zh)}</subject>")
            p.append("</subj-group>")
        if art.col_en:
            p.append('<subj-group subj-group-type="heading" xml:lang="en">')
            p.append(f"<subject>{xml_escape(art.col_en)}</subject>")
            p.append("</subj-group>")
        if art.clc:
            clc_subs = "".join(f"<subject>{xml_escape(c)}</subject>" for c in art.clc)
            p.append('<subj-group subj-group-type="clc">')
            p.append(clc_subs)
            p.append("</subj-group>")
        p.append("</article-categories>")

    # <title-group>
    if art.title_zh or art.title_en:
        p.append("<title-group>")
        if art.title_zh:
            p.append(f"<article-title>{convert_super_sub(art.title_zh)}</article-title>")
        if art.title_en:
            p.append('<trans-title-group xml:lang="en">')
            p.append(f"<trans-title>{convert_super_sub(art.title_en)}</trans-title>")
            p.append("</trans-title-group>")
        p.append("</title-group>")

    # <contrib-group>
    if art.authors_zh or art.authors_en:
        p.append('<contrib-group content-type="author">')

        n_zh = len(art.authors_zh)
        n_en = len(art.authors_en)
        n = max(n_zh, n_en)

        corresp_idx: Optional[int] = None
        if art.has_separate_corresp and art.corresp_name_zh:
            for i, a in enumerate(art.authors_zh):
                if a["name_zh"] == art.corresp_name_zh:
                    corresp_idx = i
                    break

        for i in range(n):
            zh = art.authors_zh[i] if i < n_zh else None
            en = art.authors_en[i] if i < n_en else None
            aff_ids = (zh["aff_ids"] if zh else (en["aff_ids"] if en else []))
            is_corresp = (i == corresp_idx)

            # <contrib>标签，contrib-type="author"，中文姓名用string-name标签
            attrs = ['contrib-type="author"']
            if is_corresp:
                attrs.append('corresp="yes"')
            p.append(f"<contrib {' '.join(attrs)}>")

            if zh and en:
                en_full = (en.get("surname_en", "") + " " + en.get("given_en", "")).strip()
                p.append("<name-alternatives>")
                p.append(f"<string-name>{xml_escape(zh['name_zh'])}</string-name>")
                p.append(f'<string-name xml:lang="en">{xml_escape(en_full)}</string-name>')
                p.append("</name-alternatives>")
            elif zh:
                p.append(f"<string-name>{xml_escape(zh['name_zh'])}</string-name>")
            elif en:
                en_full = (en.get("surname_en", "") + " " + en.get("given_en", "")).strip()
                p.append(f'<string-name xml:lang="en">{xml_escape(en_full)}</string-name>')

            if aff_ids:
                xrefs = "".join(
                    f'<xref ref-type="aff" rid="aff{aid}">{aid}</xref>' for aid in aff_ids
                )
                p.append(xrefs)

            if i == 0 and art.first_bio:
                bio_html = _wrap_email_in_text(convert_super_sub(art.first_bio))
                p.append(f"<bio><heading>作者简介：</heading><p>{bio_html}</p></bio>")

            if is_corresp:
                p.append('<xref ref-type="corresp" rid="COR1" />')

            if i == 0 and art.first_email:
                p.append(f"<email>{xml_escape(art.first_email)}</email>")
            elif is_corresp and art.corresp_email:
                p.append(f"<email>{xml_escape(art.corresp_email)}</email>")

            p.append("</contrib>")

        if art.aff_zh or art.aff_en:
            ids_zh = {a["id"]: a for a in art.aff_zh}
            ids_en = {a["id"]: a for a in art.aff_en}
            all_ids = sorted({x for x in ids_zh} | {x for x in ids_en},
                             key=lambda x: (x == "", x))
            for aid in all_ids:
                p.append(f'<aff-alternatives id="aff{aid}">')
                zh = ids_zh.get(aid)
                if zh:
                    label = f"<label>{xml_escape(aid)}</label>" if aid else ""
                    p.append(f"<aff>{label}{xml_escape(zh['text'])}</aff>")
                en = ids_en.get(aid)
                if en:
                    label = f"<label>{xml_escape(aid)}</label>" if aid else ""
                    p.append(f'<aff xml:lang="en">{label}{xml_escape(en["text"])}</aff>')
                p.append("</aff-alternatives>")

        p.append("</contrib-group>")

    if art.has_separate_corresp and art.corresp_bio:
        p.append("<author-notes>")
        corresp_html = _wrap_email_in_text(convert_super_sub(art.corresp_bio))
        p.append(
            f'<corresp id="COR1"><heading>通讯作者：</heading>{corresp_html}</corresp>'
        )
        p.append("</author-notes>")

    if art.pub_date:
        y, mo, d = art.pub_date.split("-")
        p.append('<pub-date date-type="pub" publication-format="print">')
        p.append(f"<day>{int(d)}</day>")
        p.append(f"<month>{int(mo)}</month>")
        p.append(f"<year>{y}</year>")
        p.append("</pub-date>")

    if art.volume:
        p.append(f"<volume>{xml_escape(art.volume)}</volume>")
    if art.issue:
        p.append(f"<issue>{art.issue}</issue>") # 两位期号，不带0
    if art.fpage:
        p.append(f"<fpage>{xml_escape(art.fpage)}</fpage>")
    if art.lpage:
        p.append(f"<lpage>{xml_escape(art.lpage)}</lpage>")

    if art.abstract_zh:
        p.append(f"<abstract><p>{convert_super_sub(art.abstract_zh)}</p></abstract>")
    if art.abstract_en:
        p.append(
            f'<abstract abstract-type="section" xml:lang="en">'
            f"<p>{convert_super_sub(art.abstract_en)}</p></abstract>"
        )

    if art.kw_zh:
        kwds = "".join(f"<kwd>{xml_escape(kw)}</kwd>" for kw in art.kw_zh)
        p.append('<kwd-group kwd-group-type="author-provided">')
        p.append(kwds)
        p.append("</kwd-group>")
    if art.kw_en:
        kwds = "".join(f"<kwd>{xml_escape(kw)}</kwd>" for kw in art.kw_en)
        p.append('<kwd-group kwd-group-type="author-provided" xml:lang="en">')
        p.append(kwds)
        p.append("</kwd-group>")

    p.extend(_build_funding_xml(art))

    p.append("</article-meta>")
    p.append("</front>")
    p.append("</article>")
    return _finalize_xml("\n".join(p))


def _finalize_xml(text: str, trailing_newline: bool = False) -> str:
    """与 example 一致: CRLF 换行; issue.xml 末尾保留空行."""
    text = text.replace("\n", "\r\n")
    if trailing_newline:
        text += "\r\n"
    return text


def build_issue_xml(
    arts: list[Article],
    year: str,
    volume: str,
    issue: str,
    journal: JournalMeta = DEFAULT_JOURNAL_META,
) -> str:
    p: list[str] = []
    p.append('\ufeff<?xml version="1.0" encoding="utf-8"?>')
    # 头部缺dtd定义补充
    p.append('<!DOCTYPE issue-xml PUBLIC " -//ZD//DTD ZD JATS Journal Archiving and Interchange Issue XML DTD v1.2 20230401//EN" "ZD-Issue-Xml.dtd">')
    p.append('<issue-xml xmlns:xlink="http://www.w3.org/1999/xlink">')

    p.append("<journal-meta>")
    p.append("<journal-title-group>")
    p.append(f"<journal-title>{xml_escape(journal.title_zh)}</journal-title>")
    p.append("</journal-title-group>")
    p.append(f'<issn publication-format="print">{xml_escape(journal.issn)}</issn>')
    p.append(f'<cn publication-format="print">{xml_escape(journal.cn)}</cn>')
    p.append("<publisher>")
    p.append(f"<publisher-name>{xml_escape(journal.publisher)}</publisher-name>")
    p.append("</publisher>")
    p.append("</journal-meta>")

    p.append("<issue-meta>")
    if arts and arts[0].pub_date:
        y, mo, d = arts[0].pub_date.split("-")
        p.append('<pub-date publication-format="print" date-type="pub">')
        p.append(f"<day>{int(d)}</day>")
        p.append(f"<month>{int(mo)}</month>")
        p.append(f"<year>{y}</year>")
        p.append("")
        p.append("</pub-date>")
    else:
        p.append('<pub-date publication-format="print" date-type="pub">')
        p.append(f"<year>{year}</year>")
        p.append("</pub-date>")
    p.append(f"<volume>{volume}</volume>")
    p.append(f"<issue>{issue}</issue>")
    p.append("</issue-meta>")

    by_col: dict = defaultdict(list)
    for art in arts:
        by_col[art.col_zh or "未分类"].append(art)

    p.append("<toc>")
    for col_name, col_arts in by_col.items():
        col_en = next((a.col_en for a in col_arts if a.col_en), "")
        col_arts.sort(key=article_sort_key)
        p.append("<issue-subject-group>")

        p.append("<issue-subject-title>")
        p.append(f"<issue-subject>{xml_escape(col_name)}</issue-subject>")
        if col_en: # 英文栏目
            p.append('<trans-issue-subject xml:lang="en">')
            p.append(f"<issue-subject>{xml_escape(col_en)}</issue-subject>")
            p.append("</trans-issue-subject>")
        p.append("</issue-subject-title><issue-article-meta>")

        for art in col_arts:
            art_fname = doi_to_filename(art.doi)
            art_href = f"..\\{art_fname}\\{art_fname}.xml"# 所有路径符号须为windows的，即\，而非/
            block = [f'<article-file xlink:href="{art_href}">']
            block.append(f"<article-title>{xml_escape(art.title_zh)}</article-title>")
            if art.fpage:
                block.append(f"<page>{xml_escape(art.fpage)}</page>")
            if art.title_en: # 英文标题
                block.append('<trans-article xml:lang="en">')
                block.append(f"<article-title>{xml_escape(art.title_en)}</article-title>")
                block.append("</trans-article>")
            block.append("</article-file>")
            p.append("\n".join(block))

        p.append("</issue-article-meta>")
        p.append("</issue-subject-group>")

    # 合并首行: <toc><issue-subject-group>
    text = "\n".join(p)
    text = text.replace("<toc>\n<issue-subject-group>", "<toc><issue-subject-group>", 1)
    p = text.split("\n")

    p.append("</toc>")
    p.append("</issue-xml>")
    return _finalize_xml("\n".join(p), trailing_newline=True)


# ============== 工具 ==============

def article_sort_key(art: Article) -> tuple:
    try:
        return (0, int(art.fpage))
    except ValueError:
        return (1, 0)


def doi_to_filename(doi: str) -> str:
    if not doi:
        return ""
    if "/" in doi:
        return doi.split("/", 1)[1]
    return doi


def article_dir_name(year: str, volume: str, issue: str, journal: JournalMeta) -> str:
    """期目录名: {期刊中文名}_{年}_{期号两位}."""
    try:
        iss = int(issue)
        return f"{journal.title_zh}_{year}_{iss:02d}"
    except ValueError:
        return f"{journal.title_zh}_{year}_{issue}"


def article_subdir_name(
    doi: str, year: str, issue: str, fpage: str, issn: str,
) -> str:
    if doi:
        return doi_to_filename(doi)
    try:
        iss = int(issue)
        return f"{issn}-{year}_{iss:02d}-{fpage}"
    except ValueError:
        return f"{issn}-{year}_{issue}-{fpage}"


def article_filename(
    doi: str, year: str, issue: str, fpage: str, issn: str,
) -> str:
    if doi:
        return f"{doi_to_filename(doi)}.xml"
    try:
        iss = int(issue)
        return f"{issn}-{year}_{iss:02d}-{fpage}.xml"
    except ValueError:
        return f"{issn}-{year}_{issue}-{fpage}.xml"


def article_index_name(year: str, issue: str, issn: str) -> str:
    try:
        iss = int(issue)
        return f"{issn}-{year}_{iss:02d}.issue.xml"
    except ValueError:
        return f"{issn}-{year}_{issue}.issue.xml"


def find_pdf(pdf_dir: Path, doi_suffix: str, path_hint: str) -> Optional[Path]:
    if doi_suffix:
        try:
            for p in pdf_dir.rglob(f"{doi_suffix}.pdf"):
                return p
        except OSError:
            pass
    if path_hint:
        cand = Path(path_hint)
        if cand.is_absolute() and cand.exists():
            return cand
        rel1 = pdf_dir / path_hint
        if rel1.exists():
            return rel1
        rel2 = pdf_dir / cand.name
        if rel2.exists():
            return rel2
    return None


# ============== Excel 载入 ==============

def load_articles(xlsx_path: Path) -> list[Article]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = list(rows[0])
    arts: list[Article] = []
    for row in rows[1:]:
        if not row or all(c is None for c in row):
            continue
        d = {header[i]: row[i] for i in range(len(header))}
        bio = parse_bio(d.get("通讯作者"))
        funding, funding_statement = parse_funding(d.get("资助"))
        a = Article(
            doi=str(d.get("doi") or "").strip(),
            title_zh=str(d.get("标题") or "").strip(),
            title_en=str(d.get("标题英") or "").strip(),
            col_zh=str(d.get("栏目") or "").strip(),
            col_en=str(d.get("栏目英") or "").strip(),
            clc=parse_clc(d.get("clc")),
            abstract_zh=str(d.get("摘要") or "").strip(),
            abstract_en=str(d.get("摘要英") or "").strip(),
            kw_zh=parse_kw(d.get("关键词")),
            kw_en=parse_kw(d.get("关键词英")),
            authors_zh=parse_authors(d.get("作者")),
            authors_en=parse_en_authors(d.get("作者英")),
            aff_zh=parse_affiliations(d.get("单位")),
            aff_en=parse_affiliations(d.get("单位英")),
            pub_date=parse_dt(d.get("日期")),
            volume=str(d.get("volume") or "").strip(),
            issue=str(d.get("issue") or "").strip(),
            fpage=str(d.get("fpage") or "").strip(),
            lpage=str(d.get("lpage") or "").strip(),
            funding=funding,
            funding_statement=funding_statement,
            pdf_path=str(d.get("path") or d.get("pdf") or "").strip(),
            first_bio=bio["first_bio"],
            first_email=bio["first_email"],
            corresp_bio=bio["corresp_bio"],
            corresp_email=bio["corresp_email"],
            has_separate_corresp=bio["has_separate_corresp"],
            corresp_name_zh=bio["corresp_name_zh"],
        )
        arts.append(a)
    return arts


# ============== 主流程 ==============

def group_articles(arts: list[Article]) -> dict[tuple, list[Article]]:
    groups: dict[tuple, list[Article]] = defaultdict(list)
    for art in arts:
        year = art.pub_date[:4] if art.pub_date else ""
        groups[(year, art.volume, art.issue)].append(art)
    for items in groups.values():
        items.sort(key=article_sort_key)
    return groups


def write_issue(
    out_root: Path,
    year: str,
    volume: str,
    issue: str,
    items: list[Article],
    journal: JournalMeta,
    pdf_dir: Optional[Path] = None,
) -> tuple[int, int]:
    """写入一期全部 XML, 可选复制 PDF. 返回 (成功 PDF 数, 失败 PDF 数)."""
    issue_dir = out_root / article_dir_name(year, volume, issue, journal)
    issue_dir.mkdir(parents=True, exist_ok=True)
    (issue_dir / "issue-files").mkdir(parents=True, exist_ok=True)

    idx_name = article_index_name(year, issue, journal.issn)
    (issue_dir / "issue-files" / idx_name).write_text(
        build_issue_xml(items, year, volume, issue, journal), encoding="utf-8"
    )

    pdf_ok = pdf_fail = 0
    for art in items:
        art_sub = article_subdir_name(art.doi, year, issue, art.fpage, journal.issn)
        art_dir = issue_dir / art_sub
        art_dir.mkdir(parents=True, exist_ok=True)
        xml_name = article_filename(art.doi, year, issue, art.fpage, journal.issn)
        (art_dir / xml_name).write_text(build_article_xml(art, journal), encoding="utf-8")

        if pdf_dir is None:
            continue
        doi_suffix = doi_to_filename(art.doi) if art.doi else ""
        src = find_pdf(pdf_dir, doi_suffix, art.pdf_path)
        if src:
            shutil.copy2(src, art_dir / xml_name.replace(".xml", ".pdf"))
            pdf_ok += 1
        else:
            pdf_fail += 1
            print(f"  [warn] PDF 未找到: {art_sub} (hint={art.pdf_path})")
    return pdf_ok, pdf_fail


def convert(
    journal: Union[JournalMeta, Mapping[str, str], None],
    xlsx_path: Union[str, Path],
    pdf_dir: Union[str, Path, None],
    output_dir: Union[str, Path],
) -> int:
    """执行转换.

    参数:
        journal:    期刊元数据 (JournalMeta 或 dict，None 则从 Excel/DOI 推断)
        xlsx_path:  维普 Excel 文件路径
        pdf_dir:    PDF 根目录 (递归查找子目录); None 则不复制 PDF
        output_dir: XML 输出目录

    返回:
        生成的 XML 文件数量; 失败时返回 0
    """
    xlsx_path = Path(xlsx_path).resolve()
    out_root = Path(output_dir).resolve()
    pdf_path = Path(pdf_dir).resolve() if pdf_dir else None

    if not xlsx_path.is_file():
        print(f"[error] Excel 不存在: {xlsx_path}", file=sys.stderr)
        return 0
    if pdf_path is not None and not pdf_path.is_dir():
        print(f"[error] PDF 目录不存在: {pdf_path}", file=sys.stderr)
        return 0

    arts = load_articles(xlsx_path)
    if not arts:
        print("[error] Excel 中无有效数据", file=sys.stderr)
        return 0

    journal_meta = build_journal_meta(xlsx_path, arts, journal)
    out_root.mkdir(parents=True, exist_ok=True)
    print(f"[info] 载入 {len(arts)} 篇文章 <- {xlsx_path}")
    print(f"[info] 期刊: {journal_meta.title_zh} | ISSN {journal_meta.issn} | CN {journal_meta.cn}")

    xml_count = 0
    for (year, volume, issue), items in group_articles(arts).items():
        pdf_ok, pdf_fail = write_issue(
            out_root, year, volume, issue, items, journal_meta, pdf_path
        )
        xml_count += len(items) + 1
        msg = f"[ok] {article_dir_name(year, volume, issue, journal_meta)}: {len(items)} 篇"
        if pdf_path is not None:
            msg += f", PDF {pdf_ok}/{pdf_ok + pdf_fail}"
        print(msg)

    print(f"\n[done] 生成 {xml_count} 个 XML -> {out_root}")
    return xml_count


# def validate_output(out_root: Path, reference: Path) -> bool:
#     """将输出与参考样例目录逐 XML 比对."""
#     ref_xmls = {p.relative_to(reference) for p in reference.rglob("*.xml")}
#     out_xmls = {p.relative_to(out_root) for p in out_root.rglob("*.xml")}
#     missing = ref_xmls - out_xmls
#     extra = out_xmls - ref_xmls
#     diffs = [
#         rel for rel in sorted(ref_xmls & out_xmls)
#         if (reference / rel).read_bytes() != (out_root / rel).read_bytes()
#     ]
#
#     ok = not missing and not extra and not diffs
#     if ok:
#         print(f"[validate] 通过: {len(ref_xmls)} 个 XML 与参考样例完全一致")
#         return True
#
#     if missing:
#         print(f"[validate] 缺少 {len(missing)} 个文件:")
#         for p in sorted(missing)[:10]:
#             print(f"  - {p}")
#     if extra:
#         print(f"[validate] 多余 {len(extra)} 个文件:")
#         for p in sorted(extra)[:10]:
#             print(f"  + {p}")
#     if diffs:
#         print(f"[validate] 内容不一致 {len(diffs)} 个文件:")
#         for p in diffs[:10]:
#             print(f"  ≠ {p}")
#     return False


def run_gui() -> None:
    """启动 Tkinter 图形界面."""
    from converter_gui import run_gui as _run
    _run()


def main():
    if "--gui" in sys.argv:
        run_gui()
        return

    ap = argparse.ArgumentParser(description="维普 Excel -> 浙大社 ZD_JATS XML")
    ap.add_argument(
        "--xlsx", default=str(DEFAULT_XLSX),
        help=f"维普数据 Excel (默认: {DEFAULT_XLSX.name})",
    )
    ap.add_argument(
        "--output-dir", default=str(DEFAULT_OUTPUT),
        help=f"输出目录 (默认: {DEFAULT_OUTPUT})",
    )
    ap.add_argument(
        "--pdf-dir", default=str(DEFAULT_PDF_DIR),
        help=f"PDF 根目录, 递归查找 (默认: {DEFAULT_PDF_DIR.name})",
    )
    ap.add_argument(
        "--with-pdf", action="store_true",
        help="同时复制 PDF (默认仅输出 XML)",
    )
    ap.add_argument("--journal-title", help="期刊中文名 (默认: 机电工程)")
    ap.add_argument("--issn", help="ISSN (默认: 1001-4551; 也可从 Excel 列或 DOI 推断)")
    ap.add_argument("--cn", help="国内刊号 CN (默认: 33-1088/TH)")
    ap.add_argument("--publisher", help="出版社 (默认: 浙江大学)")
    ap.add_argument("--gui", action="store_true", help="启动图形界面")
    # ap.add_argument(
    #     "--validate", metavar="REF_DIR",
    #     help="转换后与参考样例目录逐文件比对",
    # )
    args = ap.parse_args()

    journal = normalize_journal({
        "title_zh": args.journal_title,
        "issn": args.issn,
        "cn": args.cn,
        "publisher": args.publisher,
    })
    pdf_dir = Path(args.pdf_dir).resolve() if args.with_pdf else None

    count = convert(journal, args.xlsx, pdf_dir, args.output_dir)
    sys.exit(0 if count > 0 else 1)

    # if args.validate:
    #     ref = Path(args.validate).resolve()
    #     if not ref.is_dir():
    #         print(f"[error] 参考目录不存在: {ref}", file=sys.stderr)
    #         sys.exit(1)
    #     sys.exit(0 if validate_output(out_root, ref) else 1)


if __name__ == "__main__":
    main()
